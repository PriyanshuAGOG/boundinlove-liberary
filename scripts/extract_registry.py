#!/usr/bin/env python3
"""Convert the canonical invitation workbook into compact AI-readable JSON."""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "docs" / "reference" / "Invitation_Component_Library_2800.xlsx"
DATA = ROOT / "data"


def slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")


def rows_as_dicts(sheet, header_row: int) -> list[dict]:
    headers = [cell.value for cell in sheet[header_row]]
    records: list[dict] = []
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(value is not None for value in row):
            continue
        record = {
            str(headers[index]): value
            for index, value in enumerate(row)
            if index < len(headers) and headers[index]
        }
        records.append(record)
    return records


def compact_component(row: dict) -> dict:
    component_id = row["Component ID"]
    return {
        "id": component_id,
        "slug": slug(row["Unique Component Specification"]),
        "category": row["Major Category"],
        "family": row["Base Family"],
        "name": row["Unique Component Specification"],
        "assetType": row["Asset Type"],
        "eventFit": row["Best Event Fit"],
        "tone": row["Style Tone"],
        "motif": row["Motif"],
        "layout": row["Layout"],
        "motion": row["Motion Pattern"],
        "trigger": row["Trigger"],
        "engine": row["Primary Engine"],
        "source": row["Research Source"],
        "sourceUrl": row["Source URL"],
        "useMode": row["Use Mode"],
        "difficulty": row["Difficulty 1-5"],
        "performance": row["Performance Cost"],
        "buildSize": row["Estimated Build Size"],
        "reusability": row["Reusability 1-5"],
        "impact": row["Visual Impact 1-5"],
        "priority": row["Build Priority 1-5"],
        "dataFields": row["Required Data Fields"],
        "reducedMotion": row["Reduced-Motion Fallback"],
        "implementation": row["Implementation Direction"],
        "status": "spec",
        "implementationPath": None,
        "signature": row["Uniqueness Signature"],
    }


def main() -> None:
    if not SOURCE.exists():
        raise SystemExit(f"Missing workbook: {SOURCE}")

    workbook = load_workbook(SOURCE, read_only=True, data_only=True)
    components = [
        compact_component(row)
        for row in rows_as_dicts(workbook["COMPONENT REGISTRY"], 1)
    ]
    categories = rows_as_dicts(workbook["CATEGORY INDEX"], 4)
    sources = rows_as_dicts(workbook["SOURCE DIRECTORY"], 5)
    themes = [
        row
        for row in rows_as_dicts(workbook["THEMES & EVENTS"], 4)
        if row.get("Family")
        and row.get("Visual Vocabulary")
        and row.get("Cultural / Production Note")
    ]
    recipes = rows_as_dicts(workbook["REMIX RECIPES"], 4)

    if len(components) != 2800:
        raise SystemExit(f"Expected 2800 components, found {len(components)}")
    ids = {item["id"] for item in components}
    signatures = {item["signature"] for item in components}
    if len(ids) != len(components) or len(signatures) != len(components):
        raise SystemExit("Component ID or signature uniqueness check failed")

    DATA.mkdir(parents=True, exist_ok=True)
    payloads = {
        "components.json": components,
        "categories.json": categories,
        "sources.json": sources,
        "themes.json": themes,
        "recipes.json": recipes,
        "registry-meta.json": {
            "schemaVersion": "1.0.0",
            "componentCount": len(components),
            "categoryCount": len(categories),
            "sourceCount": len(sources),
            "themeCount": len(themes),
            "recipeCount": len(recipes),
            "canonicalWorkbook": str(SOURCE.relative_to(ROOT)),
            "statusVocabulary": ["spec", "building", "qa", "ready", "deprecated"],
        },
    }
    for filename, payload in payloads.items():
        (DATA / filename).write_text(
            json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
        )

    print(
        json.dumps(
            {
                "components": len(components),
                "categories": len(categories),
                "themes": len(themes),
                "recipes": len(recipes),
            }
        )
    )


if __name__ == "__main__":
    main()
